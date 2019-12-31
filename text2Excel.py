# !/usr/bin/env python
# coding=utf-8
'''
@Author: Dongx
@Description: 
@Date: 2019-12-27 01:37:33
@LastEditors: Dongx
@LastEditTime: 2019-12-31 14:20:52
'''
import openpyxl

# 创建excel
excel = openpyxl.Workbook()
# 创建sheet页
sheet = excel.create_sheet(index=0)

# 打开txt文档  要求utf-8编码
with open("/home/dongx/Desktop/1405客户信息1.txt", "r") as f:
	# 逐行读取
	textlist = f.readlines()
	# openpyxl要求 row col从1开始，+1保证输出最后一条数据
	for i in range(1, len(textlist) + 1):
		line = textlist[i-1]
		# 切割数据
		msg = line.split(",")
		print(line)
		# 写入每列数据
		for j in range(1, len(msg)):
			# i行j列
			sheet.cell(i, j).value = msg[j-1]
	print("for finish")
	# 保存数据
	excel.save("wangxiao2.xlsx")
	print("success")
	
